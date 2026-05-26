"""
Batch Evaluation Runner
=======================

Runs the original evaluate.py on every (run, stage) combination.
Logs successes and crashes. Then builds a master comparison Excel.

Stages evaluated:
- raw     : LLM_results CSV/         (original LLM output)
- cleaned : cleaned_data/            (old clean_csv.py output)
- ordered : formatted/ordered/       (format converter only)
- a       : formatted/stage_a/
- ade     : formatted/stage_ade/
- ab      : formatted/stage_ab/
- abde    : formatted/stage_abde/
- abc     : formatted/stage_abc/
- abcde   : formatted/stage_abcde/   (final, fully cleaned)

USAGE:
    python3 batch_evaluate.py

Output:
    LLM_evaluate/eval_<run_id>_<stage>.csv
    LLM_evaluate/_crashes.log
    LLM_evaluate/master_comparison.xlsx
"""

import os
import subprocess
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import ColorScaleRule


# Stage definitions: (stage_label, input_folder, file_suffix)
STAGES = [
    ("raw",     "LLM_results CSV",       ".csv"),
    ("cleaned", "cleaned_data",          "_cleaned.csv"),
    ("ordered", "formatted/ordered",     "_ordered.csv"),
    ("a",       "formatted/stage_a",     "_a.csv"),
    ("ade",     "formatted/stage_ade",   "_ade.csv"),
    ("ab",      "formatted/stage_ab",    "_ab.csv"),
    ("abde",    "formatted/stage_abde",  "_abde.csv"),
    ("abc",     "formatted/stage_abc",   "_abc.csv"),
    ("abcde",   "formatted/stage_abcde", "_abcde.csv"),
]

# Run IDs (16 total)
RUN_IDS = [
    f"{model}_{size}_{prompt}_run1"
    for size in ["small", "medium"]
    for model in ["llama", "qwen"]
    for prompt in ["p0", "p1", "p2", "p3"]
]

EVAL_SCRIPT = "evaluation/evaluate.py"
GSML = "gsml_30_rows.csv"
TOKEN_LOOKUP = "token_lookup.yaml"
TEXT_DIR = "texts"
OUTPUT_DIR = "LLM_evaluate"
CRASH_LOG = os.path.join(OUTPUT_DIR, "_crashes.log")


def run_evaluation(submitted_path, output_csv):
    """Run evaluate.py. Return (success, error_message)."""
    cmd = [
        "python3", EVAL_SCRIPT,
        "--gsml", GSML,
        "--submitted", submitted_path,
        "--token_lookup", TOKEN_LOOKUP,
        "--text_dir", TEXT_DIR,
        "--csv_out", output_csv,
    ]
    try:
        result = subprocess.run(cmd, capture_output=True, text=True, timeout=120)
        if result.returncode != 0:
            return False, result.stderr.strip().split("\n")[-1][:200]
        if not os.path.exists(output_csv):
            return False, "No output CSV produced"
        return True, None
    except subprocess.TimeoutExpired:
        return False, "TIMEOUT"
    except Exception as e:
        return False, str(e)[:200]


def main():
    os.makedirs(OUTPUT_DIR, exist_ok=True)
    crash_lines = []
    success_log = []

    print(f"Running {len(RUN_IDS)} runs × {len(STAGES)} stages = {len(RUN_IDS) * len(STAGES)} evaluations")
    print("=" * 80)

    for run_id in RUN_IDS:
        print(f"\n[{run_id}]")
        for stage_label, in_dir, suffix in STAGES:
            in_file = f"results_{run_id}{suffix}"
            in_path = os.path.join(in_dir, in_file)
            out_csv = os.path.join(OUTPUT_DIR, f"eval_{run_id}_{stage_label}.csv")

            if not os.path.exists(in_path):
                print(f"  {stage_label:8s}: MISSING ({in_path})")
                crash_lines.append(f"{run_id},{stage_label},MISSING,{in_path}")
                continue

            ok, err = run_evaluation(in_path, out_csv)
            if ok:
                print(f"  {stage_label:8s}: OK")
                success_log.append((run_id, stage_label, out_csv))
            else:
                print(f"  {stage_label:8s}: CRASH - {err[:80]}")
                crash_lines.append(f"{run_id},{stage_label},CRASH,{err}")

    # Write crash log
    with open(CRASH_LOG, "w") as f:
        f.write("run_id,stage,status,details\n")
        for line in crash_lines:
            f.write(line + "\n")

    print("\n" + "=" * 80)
    print(f"Successful evals: {len(success_log)}")
    print(f"Crashes/missing : {len(crash_lines)}")
    print(f"Crash log       : {CRASH_LOG}")

    # Build master comparison
    print("\nBuilding master comparison Excel...")
    build_master_excel(success_log)


def parse_run_id(run_id):
    """Parse 'qwen_medium_p1_run1' -> (model, size, prompt)."""
    parts = run_id.split("_")
    model_map = {"llama": "Llama", "qwen": "Qwen"}
    size_map = {
        ("llama", "small"): "Llama-3.1-8B",
        ("llama", "medium"): "Llama-3.1-70B",
        ("qwen", "small"): "Qwen-2.5-7B",
        ("qwen", "medium"): "Qwen-2.5-72B",
    }
    return size_map.get((parts[0], parts[1]), run_id), parts[1].capitalize(), parts[2].upper()


def fnum(x):
    try:
        v = float(x)
        return v if pd.notna(v) else None
    except (TypeError, ValueError):
        return None


def build_master_excel(success_log):
    rows = []
    for run_id, stage, csv_path in success_log:
        try:
            df = pd.read_csv(csv_path)
            combined = df[df["categories"].str.contains(r"\|", na=False)]
            if combined.empty:
                continue
            r = combined.iloc[0]
            model, size, prompt = parse_run_id(run_id)
            rows.append({
                "Model": model, "Size": size, "Prompt": prompt, "Stage": stage,
                "Recall": fnum(r.get("recall")),
                "Precision": fnum(r.get("precision")),
                "Token_Recall": fnum(r.get("token_recall")),
                "Token_Precision": fnum(r.get("token_precision")),
            })
        except Exception as e:
            print(f"  Error parsing {csv_path}: {e}")

    if not rows:
        print("No data to build Excel from.")
        return

    df = pd.DataFrame(rows)
    stage_order = ["raw", "cleaned", "ordered", "a", "ade", "ab", "abde", "abc", "abcde"]
    df["_so"] = df["Stage"].map({s: i for i, s in enumerate(stage_order)})
    size_order = {"Small": 0, "Medium": 1}
    prompt_order = {"P0": 0, "P1": 1, "P2": 2, "P3": 3}
    df["_sz"] = df["Size"].map(size_order)
    df["_pr"] = df["Prompt"].map(prompt_order)
    df = df.sort_values(["_sz", "Model", "_pr", "_so"]).drop(columns=["_so", "_sz", "_pr"]).reset_index(drop=True)

    out_path = os.path.join(OUTPUT_DIR, "master_comparison.xlsx")
    write_excel(df, out_path)
    print(f"Master comparison saved: {out_path}")


def write_excel(df, out_path):
    wb = Workbook()
    ws = wb.active
    ws.title = "Master Comparison"

    # Styles
    h_font = Font(bold=True, color="FFFFFF", size=11)
    h_fill = PatternFill("solid", start_color="2C3E50")
    title_font = Font(bold=True, size=14)
    body_font = Font(size=10)
    center = Alignment(horizontal="center", vertical="center")
    left = Alignment(horizontal="left", vertical="center")
    thin = Border(
        left=Side(style="thin", color="CCCCCC"),
        right=Side(style="thin", color="CCCCCC"),
        top=Side(style="thin", color="CCCCCC"),
        bottom=Side(style="thin", color="CCCCCC"),
    )

    # Title
    ws["A1"] = "Master Comparison — Evaluation Metrics by Stage"
    ws["A1"].font = title_font
    ws.merge_cells("A1:H1")

    ws["A2"] = ("Each row = (model, prompt, stage). Stage = which point in the post-processing pipeline. "
                "raw → cleaned → ordered → a → ab → abc → abcd → abcde.")
    ws["A2"].font = Font(size=9, italic=True, color="666666")
    ws.merge_cells("A2:H2")
    ws.row_dimensions[2].height = 30
    ws["A2"].alignment = Alignment(wrap_text=True, vertical="top")

    headers = ["Model", "Size", "Prompt", "Stage", "Recall", "Precision", "Token Recall", "Token Precision"]
    for i, h in enumerate(headers, start=1):
        c = ws.cell(row=4, column=i, value=h)
        c.font = h_font; c.fill = h_fill; c.alignment = center; c.border = thin
    ws.row_dimensions[4].height = 22

    row_idx = 5
    for _, r in df.iterrows():
        vals = [r["Model"], r["Size"], r["Prompt"], r["Stage"],
                r["Recall"], r["Precision"], r["Token_Recall"], r["Token_Precision"]]
        for col_idx, v in enumerate(vals, start=1):
            c = ws.cell(row=row_idx, column=col_idx, value=v)
            c.font = body_font; c.border = thin
            c.alignment = left if col_idx <= 4 else center
            if col_idx in (5, 6, 7, 8) and v is not None:
                c.number_format = "0.000"
        row_idx += 1

    widths = [16, 10, 10, 10, 11, 11, 14, 16]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

    last_row = row_idx - 1
    for col in ["E", "F", "G", "H"]:
        ws.conditional_formatting.add(
            f"{col}5:{col}{last_row}",
            ColorScaleRule(
                start_type="min", start_color="F8D7DA",
                mid_type="percentile", mid_value=50, mid_color="FFF3CD",
                end_type="max", end_color="D4EDDA",
            )
        )

    ws.freeze_panes = "A5"
    wb.save(out_path)


if __name__ == "__main__":
    main()