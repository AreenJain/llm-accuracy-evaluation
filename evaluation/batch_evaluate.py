"""
Batch Evaluation Runner
=======================

Runs the official scorer (evaluate.py) against every (run, stage)
combination and then builds a single Excel report comparing them all.

How it works:
  1. Discover what runs we have. We look at results/llm_csv/ and pull
     the run id out of each "results_<run_id>.csv" filename. This
     means full-story runs and sentence-mode runs are both picked up
     automatically as long as their CSVs are present.
  2. For each run, try every stage in STAGES. If the CSV for that
     stage exists, hand it to evaluate.py. If it doesn't, mark as
     MISSING. If evaluate.py errors out, mark as CRASH.
  3. Save every per-(run, stage) eval CSV to results/eval_outputs/.
  4. Read back the "combined" row from each successful eval CSV (the
     one whose `categories` column joins multiple types with '|'),
     pull recall / precision / token recall / token precision out of
     it, and write a colour-coded Excel comparing every run at every
     stage.

Stages evaluated (in order):
- raw     : results/llm_csv/                 (straight LLM output)
- ordered : results/formatted/ordered/       (format converter only)
- a       : results/formatted/stage_a/
- ade     : results/formatted/stage_ade/
- ab      : results/formatted/stage_ab/
- abde    : results/formatted/stage_abde/
- abc     : results/formatted/stage_abc/
- abcde   : results/formatted/stage_abcde/   (final, fully cleaned)

Usage:
    python3 evaluation/batch_evaluate.py

Outputs:
    results/eval_outputs/eval_<run_id>_<stage>.csv
    results/eval_outputs/_crashes.log
    results/eval_outputs/master_comparison.xlsx
"""

import os
import subprocess
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import ColorScaleRule


# Stage definitions: (stage_label, input_folder, file_suffix_to_match).
# The order here is also the visual order in the Excel report.
STAGES = [
    ("raw",     "results/llm_csv",               ".csv"),
    ("ordered", "results/formatted/ordered",     "_ordered.csv"),
    ("a",       "results/formatted/stage_a",     "_a.csv"),
    ("ade",     "results/formatted/stage_ade",   "_ade.csv"),
    ("ab",      "results/formatted/stage_ab",    "_ab.csv"),
    ("abde",    "results/formatted/stage_abde",  "_abde.csv"),
    ("abc",     "results/formatted/stage_abc",   "_abc.csv"),
    ("abcde",   "results/formatted/stage_abcde", "_abcde.csv"),
]


def _discover_run_ids(csv_dir="results/llm_csv"):
    """Pull a clean list of run ids out of the filenames in csv_dir.

    A run id is the piece between "results_" and ".csv", e.g.
    "llama_medium_p0_run1" or "llama_medium_p0_sent_run1". We skip
    any file whose stem ends in a stage suffix (someone may have
    accidentally copied processed files into the raw folder)."""
    ids = set()
    if not os.path.isdir(csv_dir):
        return []
    for f in os.listdir(csv_dir):
        if f.startswith("results_") and f.endswith(".csv"):
            stem = f[len("results_"):-len(".csv")]
            for tail in ("_cleaned", "_ordered", "_a", "_ade", "_ab", "_abde",
                         "_abc", "_abcd", "_abcde"):
                if stem.endswith(tail):
                    stem = None
                    break
            if stem:
                ids.add(stem)
    return sorted(ids)


# Computed once at import time so the list is visible in the run banner.
RUN_IDS = _discover_run_ids()

# Paths used by every invocation. EVAL_SCRIPT is run from the project root.
EVAL_SCRIPT = "evaluation/evaluate.py"
GSML = "data/gsml_30_rows.csv"
TOKEN_LOOKUP = "data/token_lookup.yaml"
TEXT_DIR = "data/texts"
OUTPUT_DIR = "results/eval_outputs"
CRASH_LOG = os.path.join(OUTPUT_DIR, "_crashes.log")


def run_evaluation(submitted_path, output_csv):
    """Invoke evaluate.py as a subprocess. Returns (ok, error_message).

    A 120-second timeout protects against the scorer hanging on
    pathological inputs."""
    cmd = [
        "python3", EVAL_SCRIPT,
        "--gsml", GSML,
        "--submitted", submitted_path,
        "--token_lookup", TOKEN_LOOKUP,
        "--text_dir", TEXT_DIR,
        "--csv_out", output_csv,
    ]
    try:
        result = subprocess.run(cmd, capture_output=True, text=True, timeout=120)
        if result.returncode != 0:
            # Keep only the last line of stderr — usually the actual error.
            return False, result.stderr.strip().split("\n")[-1][:200]
        if not os.path.exists(output_csv):
            return False, "No output CSV produced"
        return True, None
    except subprocess.TimeoutExpired:
        return False, "TIMEOUT"
    except Exception as e:
        return False, str(e)[:200]


def main():
    os.makedirs(OUTPUT_DIR, exist_ok=True)
    crash_lines = []
    success_log = []

    print(f"Running {len(RUN_IDS)} runs × {len(STAGES)} stages = {len(RUN_IDS) * len(STAGES)} evaluations")
    print("=" * 80)

    # Outer loop = each run. Inner loop = each stage of that run.
    for run_id in RUN_IDS:
        print(f"\n[{run_id}]")
        for stage_label, in_dir, suffix in STAGES:
            in_file = f"results_{run_id}{suffix}"
            in_path = os.path.join(in_dir, in_file)
            out_csv = os.path.join(OUTPUT_DIR, f"eval_{run_id}_{stage_label}.csv")

            # Skip cleanly if the upstream stage didn't produce this file.
            if not os.path.exists(in_path):
                print(f"  {stage_label:8s}: MISSING ({in_path})")
                crash_lines.append(f"{run_id},{stage_label},MISSING,{in_path}")
                continue

            ok, err = run_evaluation(in_path, out_csv)
            if ok:
                print(f"  {stage_label:8s}: OK")
                success_log.append((run_id, stage_label, out_csv))
            else:
                print(f"  {stage_label:8s}: CRASH - {err[:80]}")
                crash_lines.append(f"{run_id},{stage_label},CRASH,{err}")

    # Persist the crash log so we can audit MISSING / CRASH cases later.
    with open(CRASH_LOG, "w") as f:
        f.write("run_id,stage,status,details\n")
        for line in crash_lines:
            f.write(line + "\n")

    print("\n" + "=" * 80)
    print(f"Successful evals: {len(success_log)}")
    print(f"Crashes/missing : {len(crash_lines)}")
    print(f"Crash log       : {CRASH_LOG}")

    print("\nBuilding master comparison Excel...")
    build_master_excel(success_log)


def parse_run_id(run_id):
    """Break a run id into (model_label, size, prompt, mode).

    Examples:
      'qwen_medium_p1_run1'        -> ('Qwen-2.5-72B', 'Medium', 'P1', 'full')
      'llama_medium_p2_sent_run1'  -> ('Llama-3.1-70B', 'Medium', 'P2', 'sent')
    """
    parts = run_id.split("_")
    size_map = {
        ("llama", "small"): "Llama-3.1-8B",
        ("llama", "medium"): "Llama-3.1-70B",
        ("qwen", "small"): "Qwen-2.5-7B",
        ("qwen", "medium"): "Qwen-2.5-72B",
    }
    model_label = size_map.get((parts[0], parts[1]), run_id)
    size = parts[1].capitalize()
    prompt = parts[2].upper()
    # The "_sent" tag, if present, always sits in parts[3:].
    mode = "sent" if "sent" in parts[3:] else "full"
    return model_label, size, prompt, mode


def fnum(x):
    """Try to coerce x to a float. Return None for blanks or junk."""
    try:
        v = float(x)
        return v if pd.notna(v) else None
    except (TypeError, ValueError):
        return None


def build_master_excel(success_log):
    """Aggregate one row per successful (run, stage) eval into a single Excel.

    The evaluator outputs many rows per CSV (one per category). We
    pick the "combined" row — the one whose `categories` field joins
    multiple types with '|' — because that gives us the overall
    metrics across all mistake types."""
    rows = []
    for run_id, stage, csv_path in success_log:
        try:
            df = pd.read_csv(csv_path)
            combined = df[df["categories"].str.contains(r"\|", na=False)]
            if combined.empty:
                # No combined-category row means there were zero matches.
                continue
            r = combined.iloc[0]
            model, size, prompt, mode = parse_run_id(run_id)
            rows.append({
                "Model": model, "Size": size, "Prompt": prompt, "Mode": mode, "Stage": stage,
                "Recall": fnum(r.get("recall")),
                "Precision": fnum(r.get("precision")),
                "Token_Recall": fnum(r.get("token_recall")),
                "Token_Precision": fnum(r.get("token_precision")),
            })
        except Exception as e:
            print(f"  Error parsing {csv_path}: {e}")

    if not rows:
        print("No data to build Excel from.")
        return

    # Sort by (Size, Model, Prompt, Mode, Stage) so the Excel reads naturally.
    df = pd.DataFrame(rows)
    stage_order = ["raw", "ordered", "a", "ade", "ab", "abde", "abc", "abcde"]
    df["_so"] = df["Stage"].map({s: i for i, s in enumerate(stage_order)})
    size_order = {"Small": 0, "Medium": 1}
    prompt_order = {"P0": 0, "P1": 1, "P2": 2, "P3": 3}
    mode_order = {"full": 0, "sent": 1}
    df["_sz"] = df["Size"].map(size_order)
    df["_pr"] = df["Prompt"].map(prompt_order)
    df["_mo"] = df["Mode"].map(mode_order)
    df = (df
          .sort_values(["_sz", "Model", "_pr", "_mo", "_so"])
          .drop(columns=["_so", "_sz", "_pr", "_mo"])
          .reset_index(drop=True))

    xlsx_path = os.path.join(OUTPUT_DIR, "master_comparison.xlsx")
    write_excel(df, xlsx_path)
    print(f"Master comparison saved: {xlsx_path}")


def write_excel(df, out_path):
    """Render df to a single-sheet Excel with header styling, frozen panes,
    and a red→yellow→green colour scale across every metric column."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Master Comparison"

    # Styling shortcuts.
    h_font = Font(bold=True, color="FFFFFF", size=11)
    h_fill = PatternFill("solid", start_color="2C3E50")
    title_font = Font(bold=True, size=14)
    body_font = Font(size=10)
    center = Alignment(horizontal="center", vertical="center")
    left = Alignment(horizontal="left", vertical="center")
    thin = Border(
        left=Side(style="thin", color="CCCCCC"),
        right=Side(style="thin", color="CCCCCC"),
        top=Side(style="thin", color="CCCCCC"),
        bottom=Side(style="thin", color="CCCCCC"),
    )

    # Top title row.
    ws["A1"] = "Master Comparison — Evaluation Metrics by Stage"
    ws["A1"].font = title_font
    ws.merge_cells("A1:I1")

    # Description row right below the title.
    ws["A2"] = ("Each row = (model, prompt, mode, stage). Mode = full (full-story per LLM call) or "
                "sent (sentence-by-sentence). Stage = point in the post-processing pipeline: "
                "raw → ordered → a → ade → ab → abde → abc → abcde.")
    ws["A2"].font = Font(size=9, italic=True, color="666666")
    ws.merge_cells("A2:I2")
    ws.row_dimensions[2].height = 30
    ws["A2"].alignment = Alignment(wrap_text=True, vertical="top")

    # Column headers go on row 4 (row 3 is left blank for spacing).
    headers = ["Model", "Size", "Prompt", "Mode", "Stage",
               "Recall", "Precision", "Token Recall", "Token Precision"]
    for i, h in enumerate(headers, start=1):
        c = ws.cell(row=4, column=i, value=h)
        c.font = h_font; c.fill = h_fill; c.alignment = center; c.border = thin
    ws.row_dimensions[4].height = 22

    # Data rows.
    row_idx = 5
    for _, r in df.iterrows():
        vals = [r["Model"], r["Size"], r["Prompt"], r["Mode"], r["Stage"],
                r["Recall"], r["Precision"], r["Token_Recall"], r["Token_Precision"]]
        for col_idx, v in enumerate(vals, start=1):
            c = ws.cell(row=row_idx, column=col_idx, value=v)
            c.font = body_font; c.border = thin
            # Categorical columns left-aligned, metric columns centered.
            c.alignment = left if col_idx <= 5 else center
            # Format metric columns to 3 decimal places.
            if col_idx in (6, 7, 8, 9) and v is not None:
                c.number_format = "0.000"
        row_idx += 1

    # Column widths tuned by hand so the sheet doesn't need resizing.
    widths = [16, 10, 10, 8, 10, 11, 11, 14, 16]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

    # Heatmap colour scale for the metric columns (F..I).
    last_row = row_idx - 1
    for col in ["F", "G", "H", "I"]:
        ws.conditional_formatting.add(
            f"{col}5:{col}{last_row}",
            ColorScaleRule(
                start_type="min", start_color="F8D7DA",
                mid_type="percentile", mid_value=50, mid_color="FFF3CD",
                end_type="max", end_color="D4EDDA",
            )
        )

    # Freeze everything above the data so headers stay visible when scrolling.
    ws.freeze_panes = "A5"
    wb.save(out_path)


if __name__ == "__main__":
    main()
