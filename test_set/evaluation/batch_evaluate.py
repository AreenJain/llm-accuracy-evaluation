"""
Batch Evaluation Runner
=======================

Runs the official scorer (evaluate.py) against every (run, stage)
combination and then builds a single Excel report comparing them all.

How it works:
  1. Discover what runs we have. We look at results/llm_csv/ and pull
     the run id out of each "results_<run_id>.csv" filename. This
     means full-story runs and sentence-mode runs are both picked up
     automatically as long as their CSVs are present.
  2. For each run, try every stage in STAGES. If the CSV for that
     stage exists, hand it to evaluate.py. If it doesn't, mark as
     MISSING. If evaluate.py errors out, mark as CRASH.
  3. Save every per-(run, stage) eval CSV to results/eval_outputs/.
  4. Read back the "combined" row from each successful eval CSV (the
     one whose `categories` column joins multiple types with '|'),
     pull recall / precision / token recall / token precision out of
     it, and write a colour-coded Excel comparing every run at every
     stage.

Stages evaluated (in order):
- raw     : results/llm_csv/                 (straight LLM output)
- ordered : results/formatted/ordered/       (format converter only)
- a       : results/formatted/stage_a/
- ade     : results/formatted/stage_ade/
- ab      : results/formatted/stage_ab/
- abde    : results/formatted/stage_abde/
- abc     : results/formatted/stage_abc/
- abcde   : results/formatted/stage_abcde/   (final, fully cleaned)

Usage:
    python3 evaluation/batch_evaluate.py

Outputs:
    results/eval_outputs/eval_<run_id>_<stage>.csv
    results/eval_outputs/_crashes.log
    results/eval_outputs/master_comparison.xlsx
"""

import os
import subprocess
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import ColorScaleRule


# Stage definitions: (stage_label, input_folder, file_suffix_to_match).
# The order here is also the visual order in the Excel report.
STAGES = [
    ("raw",     "results/llm_csv",               ".csv"),
    ("ordered", "results/formatted/ordered",     "_ordered.csv"),
    ("a",       "results/formatted/stage_a",     "_a.csv"),
    ("ade",     "results/formatted/stage_ade",   "_ade.csv"),
    ("ab",      "results/formatted/stage_ab",    "_ab.csv"),
    ("abde",    "results/formatted/stage_abde",  "_abde.csv"),
    ("abc",     "results/formatted/stage_abc",   "_abc.csv"),
    ("abcde",   "results/formatted/stage_abcde", "_abcde.csv"),
]


def _discover_run_ids(csv_dir="results/llm_csv"):
    """Pull a clean list of run ids out of the filenames in csv_dir.

    A run id is the piece between "results_" and ".csv", e.g.
    "llama_medium_p0_run1" or "llama_medium_p0_sent_run1". We skip
    any file whose stem ends in a stage suffix (someone may have
    accidentally copied processed files into the raw folder)."""
    ids = set()
    if not os.path.isdir(csv_dir):
        return []
    for f in os.listdir(csv_dir):
        if f.startswith("results_") and f.endswith(".csv"):
            stem = f[len("results_"):-len(".csv")]
            for tail in ("_cleaned", "_ordered", "_a", "_ade", "_ab", "_abde",
                         "_abc", "_abcd", "_abcde"):
                if stem.endswith(tail):
                    stem = None
                    break
            if stem:
                ids.add(stem)
    return sorted(ids)


# Computed once at import time so the list is visible in the run banner.
RUN_IDS = _discover_run_ids()

# Paths used by every invocation. EVAL_SCRIPT is run from the project root.
EVAL_SCRIPT = "evaluation/evaluate.py"
GSML = "data/gsml_30_rows.csv"
TOKEN_LOOKUP = "data/token_lookup.yaml"
TEXT_DIR = "data/texts"
OUTPUT_DIR = "results/eval_outputs"
CRASH_LOG = os.path.join(OUTPUT_DIR, "_crashes.log")


def run_evaluation(submitted_path, output_csv):
    """Invoke evaluate.py as a subprocess. Returns (ok, error_message).

    A 120-second timeout protects against the scorer hanging on
    pathological inputs."""
    cmd = [
        "python3", EVAL_SCRIPT,
        "--gsml", GSML,
        "--submitted", submitted_path,
        "--token_lookup", TOKEN_LOOKUP,
        "--text_dir", TEXT_DIR,
        "--csv_out", output_csv,
    ]
    try:
        result = subprocess.run(cmd, capture_output=True, text=True, timeout=120)
        if result.returncode != 0:
            # Keep only the last line of stderr, usually the actual error.
            return False, result.stderr.strip().split("\n")[-1][:200]
        if not os.path.exists(output_csv):
            return False, "No output CSV produced"
        return True, None
    except subprocess.TimeoutExpired:
        return False, "TIMEOUT"
    except Exception as e:
        return False, str(e)[:200]


def main():
    os.makedirs(OUTPUT_DIR, exist_ok=True)
    crash_lines = []
    success_log = []

    print(f"Running {len(RUN_IDS)} runs × {len(STAGES)} stages = {len(RUN_IDS) * len(STAGES)} evaluations")
    print("=" * 80)

    # Outer loop = each run. Inner loop = each stage of that run.
    for run_id in RUN_IDS:
        print(f"\n[{run_id}]")
        for stage_label, in_dir, suffix in STAGES:
            in_file = f"results_{run_id}{suffix}"
            in_path = os.path.join(in_dir, in_file)
            out_csv = os.path.join(OUTPUT_DIR, f"eval_{run_id}_{stage_label}.csv")

            # Skip cleanly if the upstream stage didn't produce this file.
            if not os.path.exists(in_path):
                print(f"  {stage_label:8s}: MISSING ({in_path})")
                crash_lines.append(f"{run_id},{stage_label},MISSING,{in_path}")
                continue

            ok, err = run_evaluation(in_path, out_csv)
            if ok:
                print(f"  {stage_label:8s}: OK")
                success_log.append((run_id, stage_label, out_csv))
            else:
                print(f"  {stage_label:8s}: CRASH - {err[:80]}")
                crash_lines.append(f"{run_id},{stage_label},CRASH,{err}")

    # Persist the crash log so we can audit MISSING / CRASH cases later.
    with open(CRASH_LOG, "w") as f:
        f.write("run_id,stage,status,details\n")
        for line in crash_lines:
            f.write(line + "\n")

    print("\n" + "=" * 80)
    print(f"Successful evals: {len(success_log)}")
    print(f"Crashes/missing : {len(crash_lines)}")
    print(f"Crash log       : {CRASH_LOG}")

    print("\nBuilding master comparison Excel...")
    build_master_excel(success_log)


def parse_run_id(run_id):
    """Break a run id into (model_label, size, prompt, mode).

    Examples:
      'qwen_medium_p1_run1'             -> ('Qwen-2.5-72B',  'Medium', 'P1',         'full')
      'llama_medium_p2_sent_run1'       -> ('Llama-3.1-70B', 'Medium', 'P2',         'sent')
      'llama_small_p4_strat_1_run1'     -> ('Llama-3.1-8B',  'Small',  'P4_STRAT_1', 'full')
      'llama_small_p4_strat_1_sent_run1'-> ('Llama-3.1-8B',  'Small',  'P4_STRAT_1', 'sent')

    Structure: <model>_<size>_<prompt...>_[sent_]run<n>
    The prompt portion is everything between size (parts[1]) and the first
    token that is either "sent" or starts with "run".  This correctly handles
    multi-part prompt names like "p4_strat_1".
    """
    parts = run_id.split("_")
    size_map = {
        ("llama", "small"):  "Llama-3.1-8B",
        ("llama", "medium"): "Llama-3.1-70B",
        ("qwen",  "small"):  "Qwen-2.5-7B",
        ("qwen",  "medium"): "Qwen-2.5-72B",
    }
    model_label = size_map.get((parts[0], parts[1]), run_id)
    size = parts[1].capitalize()

    # Collect prompt tokens: everything after size until we hit "sent" or "run*"
    prompt_parts = []
    mode = "full"
    for tok in parts[2:]:
        if tok == "sent":
            mode = "sent"
        elif tok.startswith("run"):
            break           # run number, stop here
        else:
            prompt_parts.append(tok)

    prompt = "_".join(prompt_parts).upper()   # e.g. "P4_STRAT_1" or "P0"
    # Show a short clean label: collapse any "P4_STRAT_1" style name to just "P4".
    if prompt.startswith("P4"):
        prompt = "P4"
    return model_label, size, prompt, mode


def fnum(x):
    """Try to coerce x to a float. Return None for blanks or junk."""
    try:
        v = float(x)
        return v if pd.notna(v) else None
    except (TypeError, ValueError):
        return None


def build_master_excel(success_log):
    """Aggregate one row per successful (run, stage) eval into a single Excel.

    The evaluator outputs many rows per CSV (one per category). We
    pick the "combined" row — the one whose `categories` field joins
    multiple types with '|' — because that gives us the overall
    metrics across all mistake types."""
    rows = []
    for run_id, stage, csv_path in success_log:
        try:
            df = pd.read_csv(csv_path)
            combined = df[df["categories"].str.contains(r"\|", na=False)]
            if combined.empty:
                # No combined-category row means there were zero matches.
                continue
            r = combined.iloc[0]
            model, size, prompt, mode = parse_run_id(run_id)
            # Skip stray runs that have no prompt in their name (e.g.
            # "qwen_medium_run1") they would add empty/garbage rows.
            if not prompt:
                continue
            rows.append({
                "Model": model, "Size": size, "Prompt": prompt, "Mode": mode, "Stage": stage,
                "Recall": fnum(r.get("recall")),
                "Precision": fnum(r.get("precision")),
                "Token_Recall": fnum(r.get("token_recall")),
                "Token_Precision": fnum(r.get("token_precision")),
            })
        except Exception as e:
            print(f"  Error parsing {csv_path}: {e}")

    if not rows:
        print("No data to build Excel from.")
        return

    # Sort by (Size, Model, Prompt, Mode, Stage) so the Excel reads naturally.
    df = pd.DataFrame(rows)
    stage_order = ["raw", "ordered", "a", "ade", "ab", "abde", "abc", "abcde"]
    df["_so"] = df["Stage"].map({s: i for i, s in enumerate(stage_order)})
    size_order = {"Small": 0, "Medium": 1}
    prompt_order = {"P0": 0, "P0A": 1, "P0B": 2, "P0C": 3, "P1": 4, "P2": 5, "P3": 6, "P4": 7}
    mode_order = {"full": 0, "sent": 1}
    df["_sz"] = df["Size"].map(size_order)
    df["_pr"] = df["Prompt"].map(prompt_order)
    df["_mo"] = df["Mode"].map(mode_order)
    df = (df
          .sort_values(["_sz", "Model", "_pr", "_mo", "_so"])
          .drop(columns=["_so", "_sz", "_pr", "_mo"])
          .reset_index(drop=True))

    xlsx_path = os.path.join(OUTPUT_DIR, "master_comparison.xlsx")
    write_excel(df, xlsx_path)
    print(f"Master comparison saved: {xlsx_path}")


def write_excel(df, out_path):
    """Render df to a pivoted Excel layout matching the design spec:

    Row headers  : Prompt (col A) | Mode (col B) | Stage (col C)
    Column groups: one group per (model, size) combo, e.g. Llama/small,
                   Llama/medium, Qwen/small, Qwen/medium.
    Each group   : recall | token_recall | precision | token_precision

    Header rows  :
      Row 1 – blank | blank | blank | <model> (merged across 4 cols) | ...
      Row 2 – blank | blank | blank | <size>  (merged across 4 cols) | ...
      Row 3 – Mode  | Stage |        recall | token_recall | precision | token_precision | ...

    Data rows start at row 4.
    Prompt label is written only on the first row of each prompt group.
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Master Comparison"

    # Styling shortcuts
    h_font_white = Font(bold=True, color="FFFFFF", size=10)
    h_font_dark  = Font(bold=True, color="FFFFFF", size=10)
    h_fill_dark  = PatternFill("solid", start_color="2C3E50")
    h_fill_mid   = PatternFill("solid", start_color="4A6FA5")
    h_fill_light = PatternFill("solid", start_color="6B8CBA")
    body_font    = Font(name="Arial", size=10)
    center       = Alignment(horizontal="center", vertical="center")
    left         = Alignment(horizontal="left",   vertical="center")
    thin         = Border(
        left=Side(style="thin", color="CCCCCC"),
        right=Side(style="thin", color="CCCCCC"),
        top=Side(style="thin", color="CCCCCC"),
        bottom=Side(style="thin", color="CCCCCC"),
    )

    METRICS = ["recall", "token_recall", "precision", "token_precision"]
    METRIC_LABELS = ["recall", "token_recall", "precision", "token_precision"]

    # Determine column groups from data
    # Model order: Llama before Qwen; size order: small before medium.
    model_order = {"Llama-3.1-8B": (0, 0), "Llama-3.1-70B": (0, 1),
                   "Qwen-2.5-7B":  (1, 0), "Qwen-2.5-72B":  (1, 1)}
    size_label  = {"Llama-3.1-8B": "small", "Llama-3.1-70B": "medium",
                   "Qwen-2.5-7B":  "small", "Qwen-2.5-72B":  "medium"}
    model_name  = {"Llama-3.1-8B": "Llama", "Llama-3.1-70B": "Llama",
                   "Qwen-2.5-7B":  "Qwen",  "Qwen-2.5-72B":  "Qwen"}

    present_models = sorted(df["Model"].unique(),
                            key=lambda m: model_order.get(m, (99, 99)))

    # Fixed row structure
    # Rows 1-3 = headers; data starts at row 4.
    ROW_MODEL  = 1
    ROW_SIZE   = 2
    ROW_METRIC = 3
    DATA_START = 4

    # Fixed left columns: Prompt (A=1), Mode (B=2), Stage (C=3)
    LEFT_COLS = 3

    def col_for(model_idx, metric_idx):
        """1-based Excel column for a given model group and metric."""
        return LEFT_COLS + 1 + model_idx * len(METRICS) + metric_idx

    # Row 1: model names (merged across 4 metric cols each)
    # Rows 1-2: left three cells are empty but styled.
    for r in (ROW_MODEL, ROW_SIZE, ROW_METRIC):
        for c in range(1, LEFT_COLS + 1):
            cell = ws.cell(row=r, column=c)
            cell.fill = h_fill_dark
            cell.border = thin

    # Write the left column headers in row 3: Prompt | Mode | Stages.
    # These must line up with the data rows below (A=Prompt, B=Mode, C=Stages).
    for col_no, label in ((1, "Prompt"), (2, "Mode"), (3, "Stages")):
        c = ws.cell(row=ROW_METRIC, column=col_no, value=label)
        c.font      = h_font_white
        c.fill      = h_fill_dark
        c.alignment = center

    for mi, model in enumerate(present_models):
        start_col = col_for(mi, 0)
        end_col   = col_for(mi, len(METRICS) - 1)
        start_ltr = get_column_letter(start_col)
        end_ltr   = get_column_letter(end_col)

        # Row 1: model family name, merged
        ws.merge_cells(f"{start_ltr}{ROW_MODEL}:{end_ltr}{ROW_MODEL}")
        c1 = ws.cell(row=ROW_MODEL, column=start_col, value=model_name[model])
        c1.font = h_font_white; c1.fill = h_fill_dark; c1.alignment = center; c1.border = thin

        # Row 2: size label, merged
        ws.merge_cells(f"{start_ltr}{ROW_SIZE}:{end_ltr}{ROW_SIZE}")
        c2 = ws.cell(row=ROW_SIZE, column=start_col, value=size_label[model])
        c2.font = h_font_white; c2.fill = h_fill_mid; c2.alignment = center; c2.border = thin

        # Row 3: individual metric labels.
        for ki, label in enumerate(METRIC_LABELS):
            c3 = ws.cell(row=ROW_METRIC, column=col_for(mi, ki), value=label)
            c3.font = h_font_dark; c3.fill = h_fill_light; c3.alignment = center; c3.border = thin

    # Build row order: (prompt, mode, stage)
    stage_order  = ["raw", "ordered", "a", "ade", "ab", "abde", "abc", "abcde"]
    prompt_order = {"P0": 0, "P0A": 1, "P0B": 2, "P0C": 3, "P1": 4, "P2": 5, "P3": 6, "P4": 7}
    mode_order   = {"full": 0, "sent": 1}

    keys = (df[["Prompt", "Mode", "Stage"]]
            .drop_duplicates()
            .assign(_pr=lambda x: x["Prompt"].map(prompt_order),
                    _mo=lambda x: x["Mode"].map(mode_order),
                    _so=lambda x: x["Stage"].map({s: i for i, s in enumerate(stage_order)}))
            .sort_values(["_pr", "_mo", "_so"])
            .drop(columns=["_pr", "_mo", "_so"])
            .itertuples(index=False))

    # Index the df for fast lookup: (Prompt, Mode, Stage, Model) - row
    lookup = {(r.Prompt, r.Mode, r.Stage, r.Model): r
              for r in df.itertuples(index=False)}

    # Write data rows 
    row_idx = DATA_START
    prev_prompt = None
    prev_mode   = None

    for prompt, mode, stage in keys:
        # Prompt label only on first row of each prompt block.
        prompt_val = prompt if prompt != prev_prompt else None
        mode_val   = mode   if (prompt, mode) != (prev_prompt, prev_mode) else None
        prev_prompt, prev_mode = prompt, mode

        # Col A - Prompt
        ca = ws.cell(row=row_idx, column=1, value=prompt_val)
        ca.font = body_font; ca.border = thin; ca.alignment = left

        # Col B – Mode
        cb = ws.cell(row=row_idx, column=2, value=mode_val)
        cb.font = body_font; cb.border = thin; cb.alignment = left

        # Col C – Stage  (always shown)
        cc = ws.cell(row=row_idx, column=3, value=stage)
        cc.font = body_font; cc.border = thin; cc.alignment = left

        # Metric columns
        for mi, model in enumerate(present_models):
            rec = lookup.get((prompt, mode, stage, model))
            for ki, metric_key in enumerate(["Recall", "Token_Recall", "Precision", "Token_Precision"]):
                val = getattr(rec, metric_key, None) if rec else None
                cm = ws.cell(row=row_idx, column=col_for(mi, ki), value=val)
                cm.font = body_font; cm.border = thin; cm.alignment = center
                if val is not None:
                    cm.number_format = "0.000"

        row_idx += 1

    #  Column widths 
    ws.column_dimensions["A"].width = 7   # Prompt
    ws.column_dimensions["B"].width = 7   # Mode
    ws.column_dimensions["C"].width = 10  # Stage
    metric_width = 14
    for mi in range(len(present_models)):
        for ki in range(len(METRICS)):
            ws.column_dimensions[get_column_letter(col_for(mi, ki))].width = metric_width

    # Heatmap colour scale per metric column 
    last_row = row_idx - 1
    for mi in range(len(present_models)):
        for ki in range(len(METRICS)):
            col_ltr = get_column_letter(col_for(mi, ki))
            ws.conditional_formatting.add(
                f"{col_ltr}{DATA_START}:{col_ltr}{last_row}",
                ColorScaleRule(
                    start_type="min",        start_color="F8D7DA",
                    mid_type="percentile",   mid_value=50, mid_color="FFF3CD",
                    end_type="max",          end_color="D4EDDA",
                )
            )

    # Row heights & freeze panes
    for r in (ROW_MODEL, ROW_SIZE, ROW_METRIC):
        ws.row_dimensions[r].height = 18
    ws.freeze_panes = f"D{DATA_START}"

    wb.save(out_path)


if __name__ == "__main__":
    main()